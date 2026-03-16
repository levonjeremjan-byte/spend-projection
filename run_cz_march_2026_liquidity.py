"""
CZ Spend Projection — March 2026 — With Liquidity Split
─────────────────────────────────────────────────────────
Runs the standard projection pipeline for ALL AMs, then post-processes
Erika's "liquidity" campaigns to separate the base liquidity discount
from her AM top-up.

Context:
  A 30% menu discount runs for all providers (general liquidity).
  Erika adds 10% on top → 40% campaigns. The 30% portion should NOT
  count as AM spend. This script splits projected spend proportionally:
    AM Spend      = projected_spend × (topup / total_discount)
    Liquidity Spend = projected_spend × (liquidity_base / total_discount)

Usage:  python3 run_cz_march_2026_liquidity.py
"""

import re
import openpyxl
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from spend_projection import (
    run,
    read_historical_csv,
    read_daily_gmv,
    read_am_files,
    read_provider_gmv_from_data_drop,
    build_historical_benchmarks,
    build_fallback_benchmarks,
    compute_weekly_gmv,
    match_and_project,
    iso_week_to_dates,
    REASON_MAP,
    REASON_DISPLAY,
)

ROOT = Path(__file__).parent

CONFIG = {
    "country_code": "CZ",
    "country_name": "Czechia",
    "projection_label": "March 2026",
    "year": 2026,
    "target_weeks": [10, 11, 12, 13],
    "error_margin": 0.25,
    "am_files_dir": str(ROOT / "AM Individual Campaign Files"),
    "historical_csv": str(ROOT / "daily_campaign_spend (4).csv"),
    "gmv_csv": str(ROOT / "cz_daily_gmv.csv"),
    "output_dir": str(ROOT / "output"),
    "exclude_ams": ["Klára Bradová"],
    "exclude_spend_objectives": ["provider_campaign_locations"],
}

ERIKA_FILENAME = "Erika Šimková - CZ campaigns 2026.xlsx"
DEFAULT_LIQUIDITY_BASE = 30  # fallback when comment has no explicit split


# ─── Styles ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
AM_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIQ_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIQ_FONT = Font(name="Calibri", bold=True, color="BF8F00", size=11)
BODY_FONT = Font(name="Calibri", size=11)
SUB_FONT = Font(name="Calibri", size=10, color="595959")
BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_row(ws, row, ncol, fill=None, font=None, fmt=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if fmt and c > 1:
            cell.number_format = fmt


# ─── Liquidity Parsing ───────────────────────────────────────────────────────

def parse_liquidity_split(comment, discount_pct):
    """Extract liquidity_base and am_topup from AM COMMENTS.

    Returns (liquidity_base, am_topup) or (None, None) if not a liquidity campaign.
    """
    if not comment:
        return None, None
    comment_lower = comment.lower()
    if "liquid" not in comment_lower:
        return None, None

    # Pattern 1: explicit "X+Y" in comment (e.g. "30+10 from p")
    xy = re.search(r"(\d+)\s*\+\s*(\d+)", comment)
    if xy:
        return int(xy.group(1)), int(xy.group(2))

    # Pattern 2: explicit "N% liquidity" or "N liquidity" (e.g. "25% liquidity")
    pct = re.search(r"(\d+)\s*%?\s*liquidity", comment_lower)
    if pct:
        base = int(pct.group(1))
        topup = (discount_pct or 0) - base
        return base, max(topup, 0)

    # Pattern 3: "expansion N in <city>" (e.g. "expansion 25 in zlín")
    exp = re.search(r"expansion\s+(\d+)\s+in", comment_lower)
    if exp:
        base = int(exp.group(1))
        topup = (discount_pct or 0) - base
        return base, max(topup, 0)

    # Fallback: "liquidity" mentioned but no numbers → use default base
    if discount_pct and discount_pct > DEFAULT_LIQUIDITY_BASE:
        return DEFAULT_LIQUIDITY_BASE, discount_pct - DEFAULT_LIQUIDITY_BASE
    elif discount_pct:
        return discount_pct, 0

    return None, None


def read_erika_liquidity_tags(cfg):
    """Re-read Erika's file to extract (week, provider_id) → liquidity split mapping."""
    erika_path = Path(cfg["am_files_dir"]) / ERIKA_FILENAME
    if not erika_path.exists():
        print(f"WARNING: {erika_path} not found")
        return {}

    wb = openpyxl.load_workbook(erika_path, data_only=True)
    tags = {}  # (week, provider_id) → {liquidity_base, am_topup, comment}

    for sn in wb.sheetnames:
        if "copy" in sn.lower() or "template" in sn.lower():
            continue
        m = re.search(r"(?:Week|W)\s*(\d+)", sn, re.IGNORECASE)
        if not m:
            continue
        wk = int(m.group(1))
        if wk not in cfg["target_weeks"]:
            continue

        ws = wb[sn]
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            week_val = int(row[0])
            if week_val not in cfg["target_weeks"]:
                continue

            pid = int(row[1])
            disc = float(row[6]) if row[6] is not None else None
            comment = str(row[23]) if len(row) > 23 and row[23] is not None else ""

            liq_base, am_topup = parse_liquidity_split(comment, disc)
            if liq_base is not None:
                tags[(week_val, pid)] = {
                    "liquidity_base_pct": liq_base,
                    "am_topup_pct": am_topup,
                    "total_discount_pct": disc,
                    "comment": comment,
                }

    wb.close()
    print(f"\nLiquidity tags: found {len(tags)} (week, provider) entries for Erika")
    return tags


def enrich_with_liquidity(projected_df, liquidity_tags):
    """Add liquidity split columns to projected_df.

    For Erika's liquidity campaigns, splits the projected spend:
      am_spend_eur       = projected × (am_topup / total_discount)
      liquidity_spend_eur = projected × (liquidity_base / total_discount)
    """
    df = projected_df.copy()
    df["is_liquidity"] = False
    df["liquidity_base_pct"] = np.nan
    df["am_topup_pct"] = np.nan
    df["liquidity_comment"] = ""
    df["am_spend_eur"] = df["weekly_projected_bolt_eur"]
    df["liquidity_spend_eur"] = 0.0

    erika_mask = df["am_name"].str.contains("Erika", case=False, na=False)
    for idx in df[erika_mask].index:
        row = df.loc[idx]
        key = (row["week"], row["provider_id"])
        if key in liquidity_tags:
            tag = liquidity_tags[key]
            total_disc = tag["total_discount_pct"] or row["discount_pct"]
            if total_disc and total_disc > 0:
                liq_ratio = tag["liquidity_base_pct"] / total_disc
                am_ratio = tag["am_topup_pct"] / total_disc
            else:
                liq_ratio = 0
                am_ratio = 1

            df.at[idx, "is_liquidity"] = True
            df.at[idx, "liquidity_base_pct"] = tag["liquidity_base_pct"]
            df.at[idx, "am_topup_pct"] = tag["am_topup_pct"]
            df.at[idx, "liquidity_comment"] = tag["comment"]
            df.at[idx, "am_spend_eur"] = row["weekly_projected_bolt_eur"] * am_ratio
            df.at[idx, "liquidity_spend_eur"] = row["weekly_projected_bolt_eur"] * liq_ratio

    n_liq = df["is_liquidity"].sum()
    total_liq = df["liquidity_spend_eur"].sum()
    total_am_adj = df["am_spend_eur"].sum()
    print(f"  {n_liq} campaigns tagged as liquidity")
    print(f"  Liquidity spend separated: {total_liq:,.0f} EUR")
    print(f"  Adjusted AM spend: {total_am_adj:,.0f} EUR")
    print(f"  Original total: {df['weekly_projected_bolt_eur'].sum():,.0f} EUR")

    return df


# ─── Excel Output ────────────────────────────────────────────────────────────

def write_person_week_reason_sheet(wb, df, cfg, weekly_gmv):
    """Main breakdown: per AM, per week, per reason with liquidity split."""
    ws = wb.create_sheet("Breakdown")
    tw = cfg["target_weeks"]
    margin = cfg["error_margin"]

    ws.cell(row=1, column=1,
            value="CZ AM Spend Projection — March 2026 — Per Person / Week / Reason").font = Font(
        name="Calibri", bold=True, size=14, color="2F5496"
    )
    ws.cell(row=2, column=1,
            value="Erika's liquidity campaigns split: AM top-up vs general liquidity base").font = Font(
        name="Calibri", italic=True, size=10, color="808080"
    )

    headers = ["Account Manager", "Reason"]
    for w in tw:
        headers += [f"W{w} AM Spend", f"W{w} Liquidity"]
    headers += ["Total AM Spend", "Total Liquidity", "Total Combined",
                f"Combined + {int(margin*100)}% margin"]
    ncol = len(headers)

    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header(ws, 4, ncol)

    am_order = (
        df.groupby("am_name")["weekly_projected_bolt_eur"]
        .sum().sort_values(ascending=False).index
    )

    r = 5
    grand_am = {w: 0 for w in tw}
    grand_liq = {w: 0 for w in tw}

    for am in am_order:
        am_df = df[df["am_name"] == am]

        # AM total row
        ws.cell(row=r, column=1, value=am)
        ws.cell(row=r, column=2, value="TOTAL")
        am_total_am = 0
        am_total_liq = 0
        for wi, w in enumerate(tw):
            wk_data = am_df[am_df["week"] == w]
            am_s = wk_data["am_spend_eur"].sum()
            liq_s = wk_data["liquidity_spend_eur"].sum()
            col_am = 3 + wi * 2
            col_liq = 4 + wi * 2
            ws.cell(row=r, column=col_am, value=am_s).number_format = '#,##0'
            ws.cell(row=r, column=col_liq, value=liq_s).number_format = '#,##0'
            am_total_am += am_s
            am_total_liq += liq_s
            grand_am[w] = grand_am[w] + am_s
            grand_liq[w] = grand_liq[w] + liq_s

        base_col = 3 + len(tw) * 2
        ws.cell(row=r, column=base_col, value=am_total_am).number_format = '#,##0'
        ws.cell(row=r, column=base_col + 1, value=am_total_liq).number_format = '#,##0'
        ws.cell(row=r, column=base_col + 2, value=am_total_am + am_total_liq).number_format = '#,##0'
        ws.cell(row=r, column=base_col + 3,
                value=(am_total_am + am_total_liq) * (1 + margin)).number_format = '#,##0'
        style_row(ws, r, ncol, fill=AM_FILL, font=TOTAL_FONT)
        r += 1

        # Per-reason sub-rows
        reasons = sorted(am_df["reason_display"].unique())
        for reason in reasons:
            rd = am_df[am_df["reason_display"] == reason]
            ws.cell(row=r, column=1, value="")
            ws.cell(row=r, column=2, value=f"  {reason}").font = SUB_FONT
            reason_am = 0
            reason_liq = 0
            for wi, w in enumerate(tw):
                wk_rd = rd[rd["week"] == w]
                am_s = wk_rd["am_spend_eur"].sum()
                liq_s = wk_rd["liquidity_spend_eur"].sum()
                col_am = 3 + wi * 2
                col_liq = 4 + wi * 2
                ws.cell(row=r, column=col_am, value=am_s).number_format = '#,##0'
                ws.cell(row=r, column=col_am).font = SUB_FONT
                ws.cell(row=r, column=col_am).border = BORDER
                ws.cell(row=r, column=col_liq, value=liq_s).number_format = '#,##0'
                ws.cell(row=r, column=col_liq).font = SUB_FONT
                ws.cell(row=r, column=col_liq).border = BORDER
                reason_am += am_s
                reason_liq += liq_s

            base_col = 3 + len(tw) * 2
            ws.cell(row=r, column=base_col, value=reason_am).number_format = '#,##0'
            ws.cell(row=r, column=base_col).font = SUB_FONT
            ws.cell(row=r, column=base_col).border = BORDER
            ws.cell(row=r, column=base_col + 1, value=reason_liq).number_format = '#,##0'
            ws.cell(row=r, column=base_col + 1).font = SUB_FONT
            ws.cell(row=r, column=base_col + 1).border = BORDER
            ws.cell(row=r, column=base_col + 2, value=reason_am + reason_liq).number_format = '#,##0'
            ws.cell(row=r, column=base_col + 2).font = SUB_FONT
            ws.cell(row=r, column=base_col + 2).border = BORDER
            r += 1
        r += 1  # blank row between AMs

    # Grand total
    ws.cell(row=r, column=1, value="GRAND TOTAL")
    ws.cell(row=r, column=2, value="")
    total_am_all = 0
    total_liq_all = 0
    for wi, w in enumerate(tw):
        col_am = 3 + wi * 2
        col_liq = 4 + wi * 2
        ws.cell(row=r, column=col_am, value=grand_am[w]).number_format = '#,##0'
        ws.cell(row=r, column=col_liq, value=grand_liq[w]).number_format = '#,##0'
        total_am_all += grand_am[w]
        total_liq_all += grand_liq[w]

    base_col = 3 + len(tw) * 2
    ws.cell(row=r, column=base_col, value=total_am_all).number_format = '#,##0'
    ws.cell(row=r, column=base_col + 1, value=total_liq_all).number_format = '#,##0'
    ws.cell(row=r, column=base_col + 2, value=total_am_all + total_liq_all).number_format = '#,##0'
    ws.cell(row=r, column=base_col + 3,
            value=(total_am_all + total_liq_all) * (1 + margin)).number_format = '#,##0'
    style_row(ws, r, ncol, fill=TOTAL_FILL, font=TOTAL_FONT)

    # Liquidity-only total row
    r += 1
    ws.cell(row=r, column=1, value="LIQUIDITY ONLY (to separate)").font = LIQ_FONT
    for wi, w in enumerate(tw):
        col_liq = 4 + wi * 2
        ws.cell(row=r, column=col_liq, value=grand_liq[w]).number_format = '#,##0'
    ws.cell(row=r, column=base_col + 1, value=total_liq_all).number_format = '#,##0'
    style_row(ws, r, ncol, fill=LIQ_FILL, font=LIQ_FONT)

    # AM-only total row (after removing liquidity)
    r += 1
    ws.cell(row=r, column=1, value="AM SPEND ONLY (excl. liquidity)").font = TOTAL_FONT
    for wi, w in enumerate(tw):
        col_am = 3 + wi * 2
        ws.cell(row=r, column=col_am, value=grand_am[w]).number_format = '#,##0'
    ws.cell(row=r, column=base_col, value=total_am_all).number_format = '#,##0'
    ws.cell(row=r, column=base_col + 3, value=total_am_all * (1 + margin)).number_format = '#,##0'
    style_row(ws, r, ncol, fill=AM_FILL, font=TOTAL_FONT)

    for c in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28


def write_erika_liquidity_detail(wb, df, cfg):
    """Detailed sheet showing Erika's liquidity campaigns with split."""
    ws = wb.create_sheet("Erika Liquidity Detail")
    tw = cfg["target_weeks"]

    ws.cell(row=1, column=1,
            value="Erika's Liquidity Campaigns — Spend Split Detail").font = Font(
        name="Calibri", bold=True, size=14, color="2F5496"
    )

    erika = df[df["am_name"].str.contains("Erika", case=False, na=False)].copy()
    liq = erika[erika["is_liquidity"]].sort_values(
        ["week", "weekly_projected_bolt_eur"], ascending=[True, False]
    )

    cols = [
        ("week", "Week", 7),
        ("provider_id", "Provider ID", 12),
        ("provider_name", "Provider", 28),
        ("city", "City", 18),
        ("campaign_type", "Campaign", 16),
        ("discount_pct", "Total Disc %", 11),
        ("liquidity_base_pct", "Liquidity Base %", 14),
        ("am_topup_pct", "AM Top-up %", 11),
        ("cost_share_pct", "Provider Share %", 14),
        ("reason_display", "Reason", 22),
        ("weekly_projected_bolt_eur", "Total Projected", 16),
        ("am_spend_eur", "AM Spend (EUR)", 16),
        ("liquidity_spend_eur", "Liquidity (EUR)", 16),
        ("match_method", "Match Method", 32),
        ("liquidity_comment", "AM Comment", 40),
    ]

    for ci, (_, label, w) in enumerate(cols, 1):
        ws.cell(row=3, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = w
    style_header(ws, 3, len(cols))

    for ri, (_, row) in enumerate(liq.iterrows(), 4):
        for ci, (key, _, _) in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row[key])
            cell.font = BODY_FONT
            cell.border = BORDER
            if key in ("weekly_projected_bolt_eur", "am_spend_eur"):
                cell.number_format = '#,##0'
            elif key == "liquidity_spend_eur":
                cell.number_format = '#,##0'
                cell.fill = LIQ_FILL
            elif key in ("discount_pct", "liquidity_base_pct", "am_topup_pct"):
                cell.number_format = '0'

    # Summary at bottom
    r = 4 + len(liq) + 1
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    total_col = next(i for i, (k, _, _) in enumerate(cols, 1) if k == "weekly_projected_bolt_eur")
    am_col = next(i for i, (k, _, _) in enumerate(cols, 1) if k == "am_spend_eur")
    liq_col = next(i for i, (k, _, _) in enumerate(cols, 1) if k == "liquidity_spend_eur")

    ws.cell(row=r, column=total_col, value=liq["weekly_projected_bolt_eur"].sum()).number_format = '#,##0'
    ws.cell(row=r, column=am_col, value=liq["am_spend_eur"].sum()).number_format = '#,##0'
    ws.cell(row=r, column=liq_col, value=liq["liquidity_spend_eur"].sum()).number_format = '#,##0'
    style_row(ws, r, len(cols), fill=TOTAL_FILL, font=TOTAL_FONT)

    # Per-week summary
    r += 2
    ws.cell(row=r, column=1, value="Per-Week Summary").font = Font(
        name="Calibri", bold=True, size=12, color="2F5496"
    )
    r += 1
    sum_headers = ["Week", "# Campaigns", "Total Projected", "AM Spend", "Liquidity Spend"]
    for ci, h in enumerate(sum_headers, 1):
        ws.cell(row=r, column=ci, value=h)
    style_header(ws, r, len(sum_headers))
    r += 1
    for w in tw:
        wk = liq[liq["week"] == w]
        ws.cell(row=r, column=1, value=f"W{w}").font = BODY_FONT
        ws.cell(row=r, column=2, value=len(wk)).font = BODY_FONT
        ws.cell(row=r, column=3, value=wk["weekly_projected_bolt_eur"].sum()).number_format = '#,##0'
        ws.cell(row=r, column=4, value=wk["am_spend_eur"].sum()).number_format = '#,##0'
        ws.cell(row=r, column=5, value=wk["liquidity_spend_eur"].sum()).number_format = '#,##0'
        style_row(ws, r, len(sum_headers), font=BODY_FONT)
        r += 1


def write_campaign_detail(wb, df, cfg):
    """Full campaign detail for all AMs with liquidity split columns."""
    ws = wb.create_sheet("All Campaigns Detail")
    ws.cell(row=1, column=1,
            value="Campaign-Level Projection — All AMs").font = Font(
        name="Calibri", bold=True, size=14, color="2F5496"
    )

    cols = [
        ("am_name", "Account Manager", 22),
        ("week", "Week", 7),
        ("provider_id", "Provider ID", 12),
        ("provider_name", "Provider", 28),
        ("city", "City", 18),
        ("campaign_type", "Campaign", 16),
        ("discount_pct", "Disc %", 8),
        ("cost_share_pct", "Provider Share %", 14),
        ("reason_display", "Reason", 22),
        ("match_method", "Match Method", 32),
        ("weekly_projected_bolt_eur", "Total Projected", 16),
        ("am_spend_eur", "AM Spend (EUR)", 16),
        ("liquidity_spend_eur", "Liquidity (EUR)", 16),
        ("is_liquidity", "Is Liquidity?", 12),
    ]

    for ci, (_, label, w) in enumerate(cols, 1):
        ws.cell(row=3, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = w
    style_header(ws, 3, len(cols))

    sorted_df = df.sort_values(
        ["am_name", "week", "weekly_projected_bolt_eur"], ascending=[True, True, False]
    )
    for ri, (_, row) in enumerate(sorted_df.iterrows(), 4):
        for ci, (key, _, _) in enumerate(cols, 1):
            val = row[key]
            if key == "is_liquidity":
                val = "Yes" if val else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if key in ("weekly_projected_bolt_eur", "am_spend_eur", "liquidity_spend_eur"):
                cell.number_format = '#,##0'
            elif key == "discount_pct":
                cell.number_format = '0'
            if row.get("is_liquidity") and key == "liquidity_spend_eur":
                cell.fill = LIQ_FILL


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    cfg = CONFIG.copy()
    tw = cfg["target_weeks"]
    year = cfg["year"]
    week_dates = {w: iso_week_to_dates(year, w) for w in tw}
    cfg["_week_dates"] = week_dates

    print("=" * 70)
    print("CZ Spend Projection — March 2026 — With Liquidity Split")
    print("=" * 70)

    # Phase 1: Standard projection pipeline
    hist_df = read_historical_csv(cfg)
    gmv_df = read_daily_gmv(cfg)
    planned_df = read_am_files(cfg)
    provider_gmv = read_provider_gmv_from_data_drop(cfg)

    benchmarks = build_historical_benchmarks(hist_df, gmv_df)
    fallback = build_fallback_benchmarks(hist_df, provider_gmv, gmv_df)
    weekly_gmv = compute_weekly_gmv(gmv_df, week_dates)

    projected_df = match_and_project(planned_df, benchmarks, fallback, provider_gmv, weekly_gmv)

    # Phase 2: Tag Erika's liquidity campaigns
    print("\n── Liquidity Split ──")
    liquidity_tags = read_erika_liquidity_tags(cfg)
    enriched_df = enrich_with_liquidity(projected_df, liquidity_tags)

    # Phase 3: Print summary
    print(f"\n── Summary by AM (EUR) ──")
    am_summary = enriched_df.groupby("am_name").agg(
        total_projected=("weekly_projected_bolt_eur", "sum"),
        am_spend=("am_spend_eur", "sum"),
        liquidity_spend=("liquidity_spend_eur", "sum"),
    ).sort_values("total_projected", ascending=False)
    print(am_summary.to_string(float_format=lambda x: f"{x:,.0f}"))

    print(f"\n── Summary by Reason (EUR) ──")
    reason_summary = enriched_df.groupby("reason_display").agg(
        total_projected=("weekly_projected_bolt_eur", "sum"),
        am_spend=("am_spend_eur", "sum"),
        liquidity_spend=("liquidity_spend_eur", "sum"),
    ).sort_values("total_projected", ascending=False)
    print(reason_summary.to_string(float_format=lambda x: f"{x:,.0f}"))

    # Phase 4: Write Excel
    output_dir = Path(cfg["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "CZ - Spend Projection - March 2026 - Liquidity Split.xlsx"

    print(f"\nWriting output to {output_path}...")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    write_person_week_reason_sheet(wb_out, enriched_df, cfg, weekly_gmv)
    write_erika_liquidity_detail(wb_out, enriched_df, cfg)
    write_campaign_detail(wb_out, enriched_df, cfg)

    wb_out.save(output_path)
    print(f"  Saved: {output_path}")

    # Final numbers
    total = enriched_df["weekly_projected_bolt_eur"].sum()
    total_am = enriched_df["am_spend_eur"].sum()
    total_liq = enriched_df["liquidity_spend_eur"].sum()
    margin = cfg["error_margin"]
    total_gmv = sum(weekly_gmv.get(w, 0) for w in tw)

    print(f"\n{'=' * 70}")
    print(f"TOTAL PROJECTED BOLT SPEND:        {total:>12,.0f} EUR")
    print(f"  of which AM Spend:               {total_am:>12,.0f} EUR")
    print(f"  of which Liquidity (to separate): {total_liq:>12,.0f} EUR")
    print(f"AM SPEND + {int(margin*100)}% MARGIN:             {total_am * (1 + margin):>12,.0f} EUR")
    print(f"TOTAL GMV:                         {total_gmv:>12,.0f} EUR")
    print(f"AM SPEND AS % OF GMV:              {total_am / total_gmv * 100:>11.3f}%")
    print(f"{'=' * 70}")

    return enriched_df


if __name__ == "__main__":
    main()
