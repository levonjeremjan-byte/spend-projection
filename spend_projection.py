"""
Spend Projection Engine
─────────────────────────
Reusable across countries and months. Never edit this file for a specific run.
All country/month specifics come from the config dict passed to run().

Usage:
    from spend_projection import run
    run(config)

See run_cz_march_2026.py for an example.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import warnings
import re

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ─── Universal Maps (same across all Bolt markets) ──────────────────────────

REASON_MAP = {
    "locations": "provider_campaign_locations",
    "churn-prevention": "provider_campaign_manual_churn_prevention",
    "commission-increase": "provider_campaign_commission_increase",
    "obligations-commitments": "provider_campaign_obligations_commitments",
    "sales-benefit": "provider_campaign_sales_benefit",
    "marketing": "provider_campaign_marketing",
    "retail-growth": "provider_campaign_retail_growth",
    "retail-profitability": "provider_campaign_retail_profitability",
    "merchant-reactivation": "provider_campaign_merchant_reactivation",
}
REASON_MAP.update({v: v for v in REASON_MAP.values()})

CAMPAIGN_TYPE_MAP = {
    "Menu Discount": "item_price",
    "Item Discount": "item_price",
    "Free Full Delivery": "delivery_price",
    "Free Base Delivery": "delivery_price",
    "Free Delivery": "delivery_price",
    "Free Delivery (capped)": "delivery_price",
    "Delivery Discount": "delivery_price",
}

REASON_DISPLAY = {
    "provider_campaign_manual_churn_prevention": "churn-prevention",
    "provider_campaign_commission_increase": "commission-increase",
    "provider_campaign_obligations_commitments": "obligations-commitments",
    "provider_campaign_sales_benefit": "sales-benefit",
    "provider_campaign_marketing": "marketing",
    "provider_campaign_retail_growth": "retail-growth",
    "provider_campaign_retail_profitability": "retail-profitability",
    "provider_campaign_merchant_reactivation": "merchant-reactivation",
    "provider_campaign_locations": "locations",
}


def iso_week_to_dates(year, week_num):
    """Convert ISO year + week number to (monday, sunday) date strings."""
    monday = datetime.strptime(f"{year}-W{week_num:02d}-1", "%G-W%V-%u")
    sunday = monday + pd.Timedelta(days=6)
    return monday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")


# ─── Data Readers ────────────────────────────────────────────────────────────

def read_historical_csv(cfg):
    print("Reading historical campaign spend CSV...")
    df = pd.read_csv(cfg["historical_csv"])
    df["Cost Share on Provider"] = (
        df["Cost Share on Provider"].str.rstrip("%").astype(float) / 100
    )
    df["Bolt spend %"] = df["Bolt spend %"].str.rstrip("%").astype(float) / 100
    df["Date"] = pd.to_datetime(df["Date"])

    original = len(df)
    df = df[~df["Account Manager"].isin(cfg["exclude_ams"])]
    df = df[~df["Spend Objective"].isin(cfg["exclude_spend_objectives"])]
    print(f"  Loaded {original:,} rows, after filtering: {len(df):,}")
    return df


def read_daily_gmv(cfg):
    print("Reading daily GMV data...")
    gmv = pd.read_csv(cfg["gmv_csv"])
    gmv["date"] = pd.to_datetime(gmv["date"])
    gmv = gmv[gmv["daily_gmv_eur"] > 50000]

    trailing_avg = gmv.tail(28)["daily_gmv_eur"].mean()
    print(f"  {len(gmv)} complete days, avg daily GMV: {gmv['daily_gmv_eur'].mean():,.0f} EUR")
    print(f"  Trailing 28-day avg GMV: {trailing_avg:,.0f} EUR (used for future weeks)")

    week_dates = cfg["_week_dates"]
    last_needed = max(end for _, end in week_dates.values())
    last_date = gmv["date"].max()
    future_end = pd.to_datetime(last_needed)

    if future_end > last_date:
        future_dates = pd.date_range(start=last_date + pd.Timedelta(days=1), end=future_end)
        if len(future_dates) > 0:
            future_df = pd.DataFrame({
                "date": future_dates,
                "daily_gmv_eur": trailing_avg,
                "delivered_orders": 0,
            })
            gmv = pd.concat([gmv, future_df], ignore_index=True)
            print(f"  Extended to {gmv['date'].max().date()} with estimated GMV")

    return gmv


def _resolve_am_sources(cfg):
    """Return list of (filename, workbook) tuples from either disk or in-memory buffers."""
    if "am_file_buffers" in cfg:
        sources = []
        for name, buf in cfg["am_file_buffers"].items():
            buf.seek(0)
            wb = openpyxl.load_workbook(buf, data_only=True)
            sources.append((name, wb))
        return sources
    am_dir = Path(cfg["am_files_dir"])
    return [
        (fpath.name, openpyxl.load_workbook(fpath, data_only=True))
        for fpath in sorted(am_dir.glob("*.xlsx"))
    ]


def read_am_files(cfg):
    print("Reading AM individual campaign files...")
    target_weeks = cfg["target_weeks"]
    all_campaigns = []

    for fname, wb in _resolve_am_sources(cfg):
        am_name = fname.rsplit(".xlsx", 1)[0].split(" - ")[0].strip()
        if am_name in cfg["exclude_ams"]:
            wb.close()
            continue

        for sheet_name in wb.sheetnames:
            sn_lower = sheet_name.lower().strip()
            if "copy" in sn_lower or "template" in sn_lower:
                continue
            match = re.search(r"(?:Week|W)\s*(\d+)", sheet_name, re.IGNORECASE)
            if not match:
                continue
            week_num = int(match.group(1))
            if week_num not in target_weeks:
                continue

            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
                week_val = row[0]
                provider_id = row[1]
                campaign_type = row[5]
                if week_val is None or provider_id is None or campaign_type is None:
                    continue
                if int(week_val) not in target_weeks:
                    continue

                reason = str(row[11] or "").strip()
                reason_slug = REASON_MAP.get(reason, reason)
                if reason_slug in cfg["exclude_spend_objectives"]:
                    continue

                am_comment = ""
                if len(row) > 23 and row[23] is not None:
                    am_comment = str(row[23])

                all_campaigns.append({
                    "am_name": am_name,
                    "am_file": fname,
                    "week": int(week_val),
                    "provider_id": int(provider_id),
                    "city": row[2] or "",
                    "provider_name": row[3] or "",
                    "commitment": row[4] or "",
                    "campaign_type": campaign_type,
                    "discount_pct": float(row[6] or 0),
                    "cost_share_pct": float(row[7] or 0),
                    "users": row[8] or "All",
                    "reason_slug": reason_slug,
                    "reason_display": REASON_DISPLAY.get(reason_slug, reason_slug),
                    "bonus_type": CAMPAIGN_TYPE_MAP.get(campaign_type, "item_price"),
                    "am_comment": am_comment,
                })
        wb.close()

    df = pd.DataFrame(all_campaigns)
    print(f"  Found {len(df)} planned campaigns across {df['am_name'].nunique()} AMs")
    print(f"  Weeks: {sorted(df['week'].unique())}")
    return df


def read_provider_gmv_from_data_drop(cfg):
    print("Reading provider GMV from DATA DROP tabs...")
    provider_gmv = {}

    for fname, wb in _resolve_am_sources(cfg):
        if "DATA DROP" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["DATA DROP"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            pid = row[1]
            gmv_val = row[18]
            if pid and gmv_val and isinstance(gmv_val, (int, float)):
                provider_gmv[int(pid)] = float(gmv_val)
        wb.close()

    print(f"  Found GMV data for {len(provider_gmv)} providers")
    return provider_gmv


# ─── Benchmarks ──────────────────────────────────────────────────────────────

def build_historical_benchmarks(hist_df, gmv_df):
    print("\nBuilding historical benchmarks...")
    gmv_lookup = gmv_df.set_index("date")["daily_gmv_eur"].to_dict()
    hist_df = hist_df.copy()
    hist_df["daily_gmv"] = hist_df["Date"].map(gmv_lookup)
    hist_df["bolt_spend_pct_gmv"] = hist_df["Bolt Spend"] / hist_df["daily_gmv"]

    bm = (
        hist_df.groupby(["Provider ID", "Bonus Type"])
        .agg(
            total_bolt_spend_eur=("Bolt Spend", "sum"),
            total_bolt_spend_pct=("bolt_spend_pct_gmv", "sum"),
            unique_days=("Date", "nunique"),
            avg_cost_share=("Cost Share on Provider", "mean"),
        )
        .reset_index()
    )
    bm["avg_daily_bolt_eur"] = bm["total_bolt_spend_eur"] / bm["unique_days"]
    bm["avg_daily_bolt_pct"] = bm["total_bolt_spend_pct"] / bm["unique_days"]
    print(f"  {len(bm)} unique (Provider ID, Bonus Type) combos")
    return bm


def build_fallback_benchmarks(hist_df, provider_gmv, gmv_df):
    print("Building fallback benchmarks for new providers...")
    gmv_lookup = gmv_df.set_index("date")["daily_gmv_eur"].to_dict()
    hist_df = hist_df.copy()
    hist_df["daily_gmv"] = hist_df["Date"].map(gmv_lookup)

    pd_agg = (
        hist_df.groupby(["Provider ID", "Bonus Type"])
        .agg(total_bolt_spend=("Bolt Spend", "sum"), unique_days=("Date", "nunique"))
        .reset_index()
    )
    pd_agg["avg_daily_bolt_eur"] = pd_agg["total_bolt_spend"] / pd_agg["unique_days"]
    pd_agg["provider_monthly_gmv"] = pd_agg["Provider ID"].map(provider_gmv)
    pd_agg = pd_agg.dropna(subset=["provider_monthly_gmv"])

    if len(pd_agg) == 0:
        print("  WARNING: No provider GMV data for fallback")
        return pd.DataFrame()

    pd_agg["provider_daily_gmv"] = pd_agg["provider_monthly_gmv"] / 30
    pd_agg["bolt_spend_pct_of_own_gmv"] = (
        pd_agg["avg_daily_bolt_eur"] / pd_agg["provider_daily_gmv"]
    )
    pd_agg["gmv_quartile"] = pd.qcut(
        pd_agg["provider_monthly_gmv"], q=4,
        labels=["Q1_small", "Q2_medium", "Q3_large", "Q4_top"], duplicates="drop",
    )

    fb = (
        pd_agg.groupby(["gmv_quartile", "Bonus Type"], observed=True)
        .agg(
            avg_bolt_pct_of_own_gmv=("bolt_spend_pct_of_own_gmv", "median"),
            avg_daily_bolt_eur=("avg_daily_bolt_eur", "median"),
            provider_count=("Provider ID", "nunique"),
        )
        .reset_index()
    )
    bounds = pd_agg.groupby("gmv_quartile", observed=True)["provider_monthly_gmv"].agg(["min", "max"])
    print("  GMV quartile ranges:")
    for q, r in bounds.iterrows():
        print(f"    {q}: {r['min']:,.0f} - {r['max']:,.0f} EUR/month")
    return fb


# ─── Match & Project ─────────────────────────────────────────────────────────

def compute_weekly_gmv(gmv_df, week_dates):
    gmv_df = gmv_df.copy()
    gmv_df["date"] = pd.to_datetime(gmv_df["date"])
    weekly = {}
    for wk, (start, end) in week_dates.items():
        mask = (gmv_df["date"] >= start) & (gmv_df["date"] <= end)
        weekly[wk] = gmv_df[mask]["daily_gmv_eur"].sum()
    return weekly


def match_and_project(planned_df, benchmarks, fallback, provider_gmv, weekly_gmv):
    print("\nMatching planned campaigns to historical data...")
    results = []
    counts = {"historical": 0, "fallback": 0, "no_match": 0}

    for _, row in planned_df.iterrows():
        pid, btype, wk = row["provider_id"], row["bonus_type"], row["week"]

        bm = benchmarks[
            (benchmarks["Provider ID"] == pid) & (benchmarks["Bonus Type"] == btype)
        ]
        avg_eur = avg_pct = 0
        method = ""

        if len(bm) > 0:
            avg_eur = bm.iloc[0]["avg_daily_bolt_eur"]
            avg_pct = bm.iloc[0]["avg_daily_bolt_pct"]
            method = "historical"
            counts["historical"] += 1
        else:
            bm_any = benchmarks[benchmarks["Provider ID"] == pid]
            if len(bm_any) > 0:
                avg_eur = bm_any["avg_daily_bolt_eur"].mean()
                avg_pct = bm_any["avg_daily_bolt_pct"].mean()
                method = "historical (diff campaign type)"
                counts["historical"] += 1
            elif not fallback.empty:
                prov_gmv = provider_gmv.get(pid)
                fb = fallback[fallback["Bonus Type"] == btype]
                if prov_gmv and prov_gmv > 0 and len(fb) > 0:
                    prov_daily = prov_gmv / 30
                    avg_eur = prov_daily * fb["avg_bolt_pct_of_own_gmv"].median()
                    wgmv = weekly_gmv.get(wk, 0)
                    avg_pct = avg_eur / (wgmv / 7) if wgmv > 0 else 0
                    method = f"fallback (GMV-based, {prov_gmv:,.0f} EUR/mo)"
                elif len(fb) > 0:
                    avg_eur = fb["avg_daily_bolt_eur"].median()
                    wgmv = weekly_gmv.get(wk, 0)
                    avg_pct = avg_eur / (wgmv / 7) if wgmv > 0 else 0
                    method = "fallback (no provider GMV)"
                else:
                    avg_eur = fallback["avg_daily_bolt_eur"].median()
                    wgmv = weekly_gmv.get(wk, 0)
                    avg_pct = avg_eur / (wgmv / 7) if wgmv > 0 else 0
                    method = "fallback (overall median)"
                counts["fallback"] += 1
            else:
                method = "NO MATCH"
                counts["no_match"] += 1

        results.append({
            **row.to_dict(),
            "match_method": method,
            "avg_daily_bolt_eur": avg_eur,
            "avg_daily_bolt_pct_gmv": avg_pct,
            "weekly_projected_bolt_eur": avg_eur * 7,
            "weekly_projected_bolt_pct_gmv": avg_pct * 7,
        })

    df = pd.DataFrame(results)
    for k, v in counts.items():
        print(f"  {k}: {v}")
    print(f"  Total projected Bolt spend: {df['weekly_projected_bolt_eur'].sum():,.0f} EUR")
    return df


# ─── Week-over-Week Analysis ─────────────────────────────────────────────────

def compute_wow_changes(projected_df, target_weeks):
    """Compare consecutive weeks: which providers were added/dropped."""
    changes = []
    for i in range(1, len(target_weeks)):
        pw, cw = target_weeks[i - 1], target_weeks[i]
        prev = projected_df[projected_df["week"] == pw]
        curr = projected_df[projected_df["week"] == cw]

        prev_pids = set(prev["provider_id"].unique())
        curr_pids = set(curr["provider_id"].unique())

        for pid in curr_pids - prev_pids:
            rows = curr[curr["provider_id"] == pid]
            changes.append({
                "Transition": f"W{pw} → W{cw}",
                "Change": "New",
                "Provider ID": pid,
                "Provider": rows.iloc[0]["provider_name"],
                "AM": rows.iloc[0]["am_name"],
                "Campaigns": ", ".join(sorted(rows["campaign_type"].unique())),
                "Reasons": ", ".join(sorted(rows["reason_display"].unique())),
                "Spend Impact (EUR)": rows["weekly_projected_bolt_eur"].sum(),
            })

        for pid in prev_pids - curr_pids:
            rows = prev[prev["provider_id"] == pid]
            changes.append({
                "Transition": f"W{pw} → W{cw}",
                "Change": "Dropped",
                "Provider ID": pid,
                "Provider": rows.iloc[0]["provider_name"],
                "AM": rows.iloc[0]["am_name"],
                "Campaigns": ", ".join(sorted(rows["campaign_type"].unique())),
                "Reasons": ", ".join(sorted(rows["reason_display"].unique())),
                "Spend Impact (EUR)": -rows["weekly_projected_bolt_eur"].sum(),
            })

    if not changes:
        return pd.DataFrame()
    return pd.DataFrame(changes).sort_values(
        ["Transition", "Change", "Spend Impact (EUR)"], ascending=[True, True, False]
    )


# ─── Liquidity Split ─────────────────────────────────────────────────────────

DEFAULT_LIQUIDITY_BASE = 30  # fallback when comment mentions liquidity but no numbers


def parse_liquidity_split(comment, discount_pct):
    """Extract (liquidity_base, am_topup) from an AM comment string.

    Returns (None, None) if the campaign is not a liquidity top-up.
    """
    if not comment:
        return None, None
    comment_lower = comment.lower()
    if "liquid" not in comment_lower:
        return None, None

    xy = re.search(r"(\d+)\s*\+\s*(\d+)", comment)
    if xy:
        return int(xy.group(1)), int(xy.group(2))

    pct = re.search(r"(\d+)\s*%?\s*liquidity", comment_lower)
    if pct:
        base = int(pct.group(1))
        topup = (discount_pct or 0) - base
        return base, max(topup, 0)

    exp = re.search(r"expansion\s+(\d+)\s+in", comment_lower)
    if exp:
        base = int(exp.group(1))
        topup = (discount_pct or 0) - base
        return base, max(topup, 0)

    if discount_pct and discount_pct > DEFAULT_LIQUIDITY_BASE:
        return DEFAULT_LIQUIDITY_BASE, discount_pct - DEFAULT_LIQUIDITY_BASE
    elif discount_pct:
        return discount_pct, 0

    return None, None


def enrich_with_liquidity(projected_df):
    """Add liquidity split columns to projected_df using the am_comment field.

    For liquidity campaigns the projected spend is split proportionally:
        am_spend_eur       = projected * (am_topup / total_discount)
        liquidity_spend_eur = projected * (liquidity_base / total_discount)
    Non-liquidity campaigns keep 100% in am_spend_eur.
    """
    df = projected_df.copy()
    df["is_liquidity"] = False
    df["liquidity_base_pct"] = np.nan
    df["am_topup_pct"] = np.nan
    df["am_spend_eur"] = df["weekly_projected_bolt_eur"]
    df["liquidity_spend_eur"] = 0.0

    for idx, row in df.iterrows():
        comment = row.get("am_comment", "")
        liq_base, am_topup = parse_liquidity_split(comment, row["discount_pct"])
        if liq_base is None:
            continue

        total_disc = liq_base + am_topup
        if total_disc > 0:
            liq_ratio = liq_base / total_disc
            am_ratio = am_topup / total_disc
        else:
            liq_ratio, am_ratio = 0, 1

        df.at[idx, "is_liquidity"] = True
        df.at[idx, "liquidity_base_pct"] = liq_base
        df.at[idx, "am_topup_pct"] = am_topup
        df.at[idx, "am_spend_eur"] = row["weekly_projected_bolt_eur"] * am_ratio
        df.at[idx, "liquidity_spend_eur"] = row["weekly_projected_bolt_eur"] * liq_ratio

    n_liq = df["is_liquidity"].sum()
    total_liq = df["liquidity_spend_eur"].sum()
    print(f"  Liquidity enrichment: {n_liq} campaigns tagged, "
          f"{total_liq:,.0f} EUR separated")
    return df


# ─── Excel Writer ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
MARGIN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
MARGIN_FONT = Font(name="Calibri", bold=True, color="BF8F00", size=11)
AM_ROW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FALLBACK_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=11)
SUB_FONT = Font(name="Calibri", size=10, color="595959")
BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def _style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def _style_row(ws, row, ncol, fill=None, font=None, fmt=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if fmt and c > 1:
            cell.number_format = fmt


def write_summary_sheet(wb, cfg, projected_df, weekly_gmv):
    ws = wb.create_sheet("Summary")
    tw = cfg["target_weeks"]
    margin = cfg["error_margin"]
    title = f"{cfg['country_code']} - Spend Projection - {cfg['projection_label']}"
    total_gmv = sum(weekly_gmv.get(w, 0) for w in tw)

    ws.cell(row=1, column=1, value=title).font = Font(
        name="Calibri", bold=True, size=14, color="2F5496"
    )

    pivot = projected_df.pivot_table(
        index="reason_display", columns="week",
        values="weekly_projected_bolt_eur", aggfunc="sum", fill_value=0,
    )
    for w in tw:
        if w not in pivot.columns:
            pivot[w] = 0
    pivot = pivot[tw]

    ncol = len(tw) + 3
    for section, is_pct in [("Projected Bolt Spend (EUR)", False), ("Projected Bolt Spend (% of GMV)", True)]:
        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=1, value=section).font = Font(
            name="Calibri", bold=True, size=12, color="2F5496"
        )

        hdr_row = start_row + 1
        headers = ["Spend Objective"] + [f"W{w}" for w in tw] + ["Total", f"Total + {int(margin*100)}% margin"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=hdr_row, column=ci, value=h)
        _style_header(ws, hdr_row, ncol)

        r = hdr_row + 1
        fmt = '0.000%' if is_pct else '#,##0'
        for reason in sorted(pivot.index):
            ws.cell(row=r, column=1, value=reason).font = BODY_FONT
            row_total = 0
            for ci, w in enumerate(tw, 2):
                val = pivot.loc[reason, w]
                if is_pct:
                    wgmv = weekly_gmv.get(w, 1)
                    val = val / wgmv if wgmv > 0 else 0
                ws.cell(row=r, column=ci, value=val).number_format = fmt
                ws.cell(row=r, column=ci).font = BODY_FONT
                ws.cell(row=r, column=ci).border = BORDER
                row_total += pivot.loc[reason, w]
            t = row_total / total_gmv if (is_pct and total_gmv > 0) else row_total
            ws.cell(row=r, column=len(tw) + 2, value=t).number_format = fmt
            ws.cell(row=r, column=len(tw) + 3, value=t * (1 + margin)).number_format = fmt
            _style_row(ws, r, ncol, font=BODY_FONT)
            r += 1

        ws.cell(row=r, column=1, value="Total")
        grand = 0
        for ci, w in enumerate(tw, 2):
            val = pivot[w].sum()
            if is_pct:
                wgmv = weekly_gmv.get(w, 1)
                val = val / wgmv if wgmv > 0 else 0
            ws.cell(row=r, column=ci, value=val).number_format = fmt
            grand += pivot[w].sum()
        t = grand / total_gmv if (is_pct and total_gmv > 0) else grand
        ws.cell(row=r, column=len(tw) + 2, value=t).number_format = fmt
        ws.cell(row=r, column=len(tw) + 3, value=t * (1 + margin)).number_format = fmt
        _style_row(ws, r, ncol, fill=TOTAL_FILL, font=TOTAL_FONT)
        r += 1

        if not is_pct:
            ws.cell(row=r, column=1, value=f"Total + {int(margin*100)}% error margin")
            for ci, w in enumerate(tw, 2):
                ws.cell(row=r, column=ci, value=pivot[w].sum() * (1 + margin)).number_format = fmt
            _style_row(ws, r, ncol, fill=MARGIN_FILL, font=MARGIN_FONT)

    for c in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions["A"].width = 28


def write_am_sheet(wb, cfg, projected_df, weekly_gmv):
    ws = wb.create_sheet("Summary per AM")
    tw = cfg["target_weeks"]
    margin = cfg["error_margin"]
    title = f"{cfg['country_code']} - Spend Projection by AM - {cfg['projection_label']}"
    total_gmv = sum(weekly_gmv.get(w, 0) for w in tw)

    ws.cell(row=1, column=1, value=title).font = Font(
        name="Calibri", bold=True, size=14, color="2F5496"
    )

    headers = ["Account Manager"] + [f"W{w}" for w in tw] + ["Total EUR", f"Total + {int(margin*100)}% margin"]
    ncol = len(headers)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header(ws, 3, ncol)

    am_order = (
        projected_df.groupby("am_name")["weekly_projected_bolt_eur"]
        .sum().sort_values(ascending=False).index
    )

    r = 4
    for am in am_order:
        am_df = projected_df[projected_df["am_name"] == am]
        wk_sums = am_df.groupby("week")["weekly_projected_bolt_eur"].sum()
        total_am = sum(wk_sums.get(w, 0) for w in tw)

        ws.cell(row=r, column=1, value=am)
        for ci, w in enumerate(tw, 2):
            ws.cell(row=r, column=ci, value=wk_sums.get(w, 0)).number_format = '#,##0'
        ws.cell(row=r, column=len(tw) + 2, value=total_am).number_format = '#,##0'
        ws.cell(row=r, column=len(tw) + 3, value=total_am * (1 + margin)).number_format = '#,##0'
        _style_row(ws, r, ncol, fill=AM_ROW_FILL, font=TOTAL_FONT)
        r += 1

        reason_wk = am_df.groupby(["reason_display", "week"])["weekly_projected_bolt_eur"].sum().unstack(fill_value=0)
        for reason in sorted(reason_wk.index):
            ws.cell(row=r, column=1, value=f"  {reason}").font = SUB_FONT
            reason_tot = 0
            for ci, w in enumerate(tw, 2):
                v = reason_wk.loc[reason].get(w, 0)
                ws.cell(row=r, column=ci, value=v).number_format = '#,##0'
                ws.cell(row=r, column=ci).font = SUB_FONT
                ws.cell(row=r, column=ci).border = BORDER
                reason_tot += v
            ws.cell(row=r, column=len(tw) + 2, value=reason_tot).number_format = '#,##0'
            ws.cell(row=r, column=len(tw) + 2).font = SUB_FONT
            ws.cell(row=r, column=len(tw) + 2).border = BORDER
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="GRAND TOTAL")
    grand_wk = projected_df.groupby("week")["weekly_projected_bolt_eur"].sum()
    grand = sum(grand_wk.get(w, 0) for w in tw)
    for ci, w in enumerate(tw, 2):
        ws.cell(row=r, column=ci, value=grand_wk.get(w, 0)).number_format = '#,##0'
    ws.cell(row=r, column=len(tw) + 2, value=grand).number_format = '#,##0'
    ws.cell(row=r, column=len(tw) + 3, value=grand * (1 + margin)).number_format = '#,##0'
    _style_row(ws, r, ncol, fill=TOTAL_FILL, font=TOTAL_FONT)

    r += 1
    ws.cell(row=r, column=1, value="% of GMV").font = Font(
        name="Calibri", italic=True, size=10, color="808080"
    )
    for ci, w in enumerate(tw, 2):
        wgmv = weekly_gmv.get(w, 1)
        ws.cell(row=r, column=ci, value=grand_wk.get(w, 0) / wgmv if wgmv > 0 else 0).number_format = '0.000%'
    ws.cell(row=r, column=len(tw) + 2,
            value=grand / total_gmv if total_gmv > 0 else 0).number_format = '0.000%'

    for c in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions["A"].width = 30


def write_detail_sheet(wb, cfg, projected_df):
    ws = wb.create_sheet("Campaign Detail")
    ws.cell(row=1, column=1, value="Campaign-Level Projection Detail").font = Font(
        name="Calibri", bold=True, size=14, color="2F5496"
    )

    cols = [
        ("am_name", "Account Manager", 22),
        ("week", "Week", 7),
        ("provider_id", "Provider ID", 12),
        ("provider_name", "Provider Name", 30),
        ("city", "City", 20),
        ("campaign_type", "Campaign Type", 16),
        ("discount_pct", "Discount %", 10),
        ("cost_share_pct", "Cost Share % (Provider)", 18),
        ("reason_display", "Spend Objective", 20),
        ("commitment", "Commitment", 12),
        ("match_method", "Match Method", 35),
        ("avg_daily_bolt_eur", "Avg Daily Bolt Spend (EUR)", 22),
        ("weekly_projected_bolt_eur", "Weekly Projected (EUR)", 22),
    ]

    for ci, (_, label, w) in enumerate(cols, 1):
        ws.cell(row=3, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = w
    _style_header(ws, 3, len(cols))

    sorted_df = projected_df.sort_values(
        ["am_name", "week", "weekly_projected_bolt_eur"], ascending=[True, True, False]
    )
    for ri, (_, row) in enumerate(sorted_df.iterrows(), 4):
        is_fb = "fallback" in str(row["match_method"]).lower() or "NO MATCH" in str(row["match_method"])
        for ci, (key, _, _) in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row[key])
            cell.font = BODY_FONT
            cell.border = BORDER
            if key in ("avg_daily_bolt_eur", "weekly_projected_bolt_eur"):
                cell.number_format = '#,##0.00'
            elif key == "discount_pct":
                cell.number_format = '0'
            if is_fb:
                cell.fill = FALLBACK_FILL


# ─── Main Runner ─────────────────────────────────────────────────────────────

def run(cfg):
    """Run the full spend projection pipeline with the given config."""
    tw = cfg["target_weeks"]
    year = cfg["year"]
    week_dates = {w: iso_week_to_dates(year, w) for w in tw}
    cfg["_week_dates"] = week_dates

    label = f"{cfg['country_code']} - {cfg['projection_label']}"
    print("=" * 60)
    print(f"Spend Projection: {label}")
    print("=" * 60)

    hist_df = read_historical_csv(cfg)
    gmv_df = read_daily_gmv(cfg)
    planned_df = read_am_files(cfg)
    provider_gmv = read_provider_gmv_from_data_drop(cfg)

    benchmarks = build_historical_benchmarks(hist_df, gmv_df)
    fallback = build_fallback_benchmarks(hist_df, provider_gmv, gmv_df)
    weekly_gmv = compute_weekly_gmv(gmv_df, week_dates)

    projected_df = match_and_project(planned_df, benchmarks, fallback, provider_gmv, weekly_gmv)

    print(f"\n── Summary by Spend Objective (EUR) ──")
    pivot = projected_df.pivot_table(
        index="reason_display", columns="week",
        values="weekly_projected_bolt_eur", aggfunc="sum", fill_value=0,
    )
    for w in tw:
        if w not in pivot.columns:
            pivot[w] = 0
    pivot["Total"] = pivot[tw].sum(axis=1)
    print(pivot[tw + ["Total"]].to_string(float_format=lambda x: f"{x:,.0f}"))

    print(f"\n── Summary by AM (EUR) ──")
    am_sum = projected_df.groupby("am_name")["weekly_projected_bolt_eur"].sum().sort_values(ascending=False)
    print(am_sum.to_string(float_format=lambda x: f"{x:,.0f}"))

    output_dir = Path(cfg.get("output_dir", "."))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{cfg['country_code']} - Spend Projection - {cfg['projection_label']}.xlsx"

    print(f"\nWriting output to {output_path}...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, cfg, projected_df, weekly_gmv)
    write_am_sheet(wb, cfg, projected_df, weekly_gmv)
    write_detail_sheet(wb, cfg, projected_df)
    wb.save(output_path)
    print(f"  Done!")

    total_spend = projected_df["weekly_projected_bolt_eur"].sum()
    total_gmv = sum(weekly_gmv.get(w, 0) for w in tw)
    margin = cfg["error_margin"]
    print(f"\n{'=' * 60}")
    print(f"TOTAL PROJECTED BOLT SPEND: {total_spend:,.0f} EUR")
    print(f"TOTAL PROJECTED GMV: {total_gmv:,.0f} EUR")
    print(f"SPEND AS % OF GMV: {total_spend / total_gmv * 100:.3f}%")
    print(f"SPEND + {int(margin*100)}% MARGIN: {total_spend * (1 + margin):,.0f} EUR")
    print(f"{'=' * 60}")

    return projected_df
