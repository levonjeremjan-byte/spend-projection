"""
CZ Spend Projection Tool - March 2026
Reads AM individual campaign files + historical daily campaign spend,
projects weekly Bolt spend per AM and per Spend Objective.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import warnings
import re

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ─── Configuration ───────────────────────────────────────────────────────────

WORKSPACE = Path("/Users/levonjeremjan/Desktop/Cursor/Cusror Projects/Spend Projection")
AM_FILES_DIR = WORKSPACE / "AM Individual Campaign Files"
HISTORICAL_CSV = WORKSPACE / "daily_campaign_spend (4).csv"
GMV_CSV = WORKSPACE / "cz_daily_gmv.csv"
OUTPUT_FILE = WORKSPACE / "CZ - Spend Projection - March 2026.xlsx"

TARGET_WEEKS = [10, 11, 12, 13]
ERROR_MARGIN = 0.25

EXCLUDE_AMS = ["Klára Bradová"]
EXCLUDE_SPEND_OBJECTIVES = ["provider_campaign_locations"]

# AM file shorthand → full slug
REASON_MAP = {
    "locations": "provider_campaign_locations",
    "churn-prevention": "provider_campaign_manual_churn_prevention",
    "commission-increase": "provider_campaign_commission_increase",
    "obligations-commitments": "provider_campaign_obligations_commitments",
    "sales-benefit": "provider_campaign_sales_benefit",
    "marketing": "provider_campaign_marketing",
    "retail-growth": "provider_campaign_retail_growth",
    "retail-profitability": "provider_campaign_retail_profitability",
    "merchant-reactivation": "provider_campaign_merchant_reactivation",
    "provider_campaign_locations": "provider_campaign_locations",
    "provider_campaign_manual_churn_prevention": "provider_campaign_manual_churn_prevention",
    "provider_campaign_commission_increase": "provider_campaign_commission_increase",
    "provider_campaign_obligations_commitments": "provider_campaign_obligations_commitments",
    "provider_campaign_sales_benefit": "provider_campaign_sales_benefit",
    "provider_campaign_marketing": "provider_campaign_marketing",
    "provider_campaign_retail_growth": "provider_campaign_retail_growth",
    "provider_campaign_retail_profitability": "provider_campaign_retail_profitability",
    "provider_campaign_merchant_reactivation": "provider_campaign_merchant_reactivation",
}

# AM file campaign type → CSV bonus_type
CAMPAIGN_TYPE_MAP = {
    "Menu Discount": "item_price",
    "Item Discount": "item_price",
    "Free Full Delivery": "delivery_price",
    "Free Base Delivery": "delivery_price",
    "Free Delivery": "delivery_price",
    "Free Delivery (capped)": "delivery_price",
    "Delivery Discount": "delivery_price",
}

REASON_DISPLAY = {
    "provider_campaign_manual_churn_prevention": "churn-prevention",
    "provider_campaign_commission_increase": "commission-increase",
    "provider_campaign_obligations_commitments": "obligations-commitments",
    "provider_campaign_sales_benefit": "sales-benefit",
    "provider_campaign_marketing": "marketing",
    "provider_campaign_retail_growth": "retail-growth",
    "provider_campaign_retail_profitability": "retail-profitability",
    "provider_campaign_merchant_reactivation": "merchant-reactivation",
    "provider_campaign_locations": "locations",
}

# Week → date ranges for March 2026 (ISO weeks)
WEEK_DATES = {
    10: ("2026-03-02", "2026-03-08"),
    11: ("2026-03-09", "2026-03-15"),
    12: ("2026-03-16", "2026-03-22"),
    13: ("2026-03-23", "2026-03-29"),
}


# ─── 1. Read Inputs ─────────────────────────────────────────────────────────

def read_historical_csv():
    print("Reading historical campaign spend CSV...")
    df = pd.read_csv(HISTORICAL_CSV)
    df["Cost Share on Provider"] = (
        df["Cost Share on Provider"].str.rstrip("%").astype(float) / 100
    )
    df["Bolt spend %"] = df["Bolt spend %"].str.rstrip("%").astype(float) / 100
    df["Date"] = pd.to_datetime(df["Date"])

    original = len(df)
    df = df[~df["Account Manager"].isin(EXCLUDE_AMS)]
    df = df[~df["Spend Objective"].isin(EXCLUDE_SPEND_OBJECTIVES)]
    print(f"  Loaded {original:,} rows, after filtering: {len(df):,}")
    return df


def read_daily_gmv():
    print("Reading daily GMV data...")
    gmv = pd.read_csv(GMV_CSV)
    gmv["date"] = pd.to_datetime(gmv["date"])

    # Drop partial days (today might be incomplete)
    gmv = gmv[gmv["daily_gmv_eur"] > 50000]

    # Estimate future daily GMV using trailing 28-day average
    trailing_avg = gmv.tail(28)["daily_gmv_eur"].mean()
    print(f"  {len(gmv)} complete days, avg daily GMV: {gmv['daily_gmv_eur'].mean():,.0f} EUR")
    print(f"  Trailing 28-day avg GMV: {trailing_avg:,.0f} EUR (used for future weeks)")

    # Fill in future dates through end of March
    last_date = gmv["date"].max()
    future_dates = pd.date_range(start=last_date + pd.Timedelta(days=1), end="2026-03-31")
    if len(future_dates) > 0:
        future_df = pd.DataFrame({
            "date": future_dates,
            "daily_gmv_eur": trailing_avg,
            "delivered_orders": 0,
        })
        gmv = pd.concat([gmv, future_df], ignore_index=True)
        print(f"  Extended to {gmv['date'].max().date()} with estimated GMV")

    return gmv


def read_am_files():
    print("Reading AM individual campaign files...")
    all_campaigns = []

    for fpath in sorted(AM_FILES_DIR.glob("*.xlsx")):
        am_name = fpath.stem.split(" - ")[0].strip()
        if am_name in EXCLUDE_AMS:
            continue

        wb = openpyxl.load_workbook(fpath, data_only=True)

        for sheet_name in wb.sheetnames:
            if "week" not in sheet_name.lower():
                continue
            match = re.search(r"Week\s+(\d+)", sheet_name, re.IGNORECASE)
            if not match:
                continue
            week_num = int(match.group(1))
            if week_num not in TARGET_WEEKS:
                continue

            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
                week_val = row[0]
                provider_id = row[1]
                campaign_type = row[5]
                if week_val is None or provider_id is None or campaign_type is None:
                    continue
                if int(week_val) not in TARGET_WEEKS:
                    continue

                city = row[2] or ""
                provider_name = row[3] or ""
                commitment = row[4] or ""
                discount_pct = row[6] or 0
                cost_share_pct = row[7] or 0
                users = row[8] or "All"
                reason = row[11] or ""

                reason_slug = REASON_MAP.get(str(reason).strip(), str(reason).strip())

                if reason_slug in EXCLUDE_SPEND_OBJECTIVES:
                    continue

                bonus_type = CAMPAIGN_TYPE_MAP.get(campaign_type, "item_price")

                all_campaigns.append({
                    "am_name": am_name,
                    "am_file": fpath.name,
                    "week": int(week_val),
                    "provider_id": int(provider_id),
                    "city": city,
                    "provider_name": provider_name,
                    "commitment": commitment,
                    "campaign_type": campaign_type,
                    "discount_pct": float(discount_pct) if discount_pct else 0,
                    "cost_share_pct": float(cost_share_pct) if cost_share_pct else 0,
                    "users": users,
                    "reason_slug": reason_slug,
                    "reason_display": REASON_DISPLAY.get(reason_slug, reason_slug),
                    "bonus_type": bonus_type,
                })

        wb.close()

    df = pd.DataFrame(all_campaigns)
    print(f"  Found {len(df)} planned campaigns across {df['am_name'].nunique()} AMs")
    print(f"  Weeks: {sorted(df['week'].unique())}")
    return df


def read_provider_gmv_from_data_drop():
    """Read provider-level GMV from the DATA DROP tab for fallback matching."""
    print("Reading provider GMV from DATA DROP tabs...")
    provider_gmv = {}

    for fpath in sorted(AM_FILES_DIR.glob("*.xlsx")):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        if "DATA DROP" not in wb.sheetnames:
            wb.close()
            continue

        ws = wb["DATA DROP"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            provider_id = row[1]
            gmv_val = row[18]  # column S = Total GMV Before Discounts
            if provider_id and gmv_val and isinstance(gmv_val, (int, float)):
                provider_gmv[int(provider_id)] = float(gmv_val)

        wb.close()

    print(f"  Found GMV data for {len(provider_gmv)} providers")
    return provider_gmv


# ─── 2. Build Historical Benchmarks ─────────────────────────────────────────

def build_historical_benchmarks(hist_df, gmv_df):
    print("\nBuilding historical benchmarks...")

    gmv_lookup = gmv_df.set_index("date")["daily_gmv_eur"].to_dict()
    hist_df["daily_gmv"] = hist_df["Date"].map(gmv_lookup)
    hist_df["bolt_spend_pct_gmv"] = hist_df["Bolt Spend"] / hist_df["daily_gmv"]

    benchmarks = (
        hist_df.groupby(["Provider ID", "Bonus Type"])
        .agg(
            total_bolt_spend_eur=("Bolt Spend", "sum"),
            total_bolt_spend_pct=("bolt_spend_pct_gmv", "sum"),
            unique_days=("Date", "nunique"),
            avg_cost_share=("Cost Share on Provider", "mean"),
        )
        .reset_index()
    )

    benchmarks["avg_daily_bolt_eur"] = (
        benchmarks["total_bolt_spend_eur"] / benchmarks["unique_days"]
    )
    benchmarks["avg_daily_bolt_pct"] = (
        benchmarks["total_bolt_spend_pct"] / benchmarks["unique_days"]
    )

    print(f"  {len(benchmarks)} unique (Provider ID, Bonus Type) combos")

    return benchmarks


def build_fallback_benchmarks(hist_df, provider_gmv, gmv_df):
    """
    For new providers: group historical providers by GMV quartile + campaign type.
    Calculate average daily Bolt spend as % of the provider's own GMV.
    """
    print("Building fallback benchmarks for new providers...")

    gmv_lookup = gmv_df.set_index("date")["daily_gmv_eur"].to_dict()
    hist_df = hist_df.copy()
    hist_df["daily_gmv"] = hist_df["Date"].map(gmv_lookup)

    provider_data = (
        hist_df.groupby(["Provider ID", "Bonus Type"])
        .agg(
            total_bolt_spend=("Bolt Spend", "sum"),
            unique_days=("Date", "nunique"),
        )
        .reset_index()
    )

    provider_data["avg_daily_bolt_eur"] = (
        provider_data["total_bolt_spend"] / provider_data["unique_days"]
    )

    provider_data["provider_monthly_gmv"] = provider_data["Provider ID"].map(provider_gmv)
    provider_data = provider_data.dropna(subset=["provider_monthly_gmv"])

    if len(provider_data) == 0:
        print("  WARNING: No provider GMV data for fallback")
        return pd.DataFrame()

    provider_data["provider_daily_gmv"] = provider_data["provider_monthly_gmv"] / 30
    provider_data["bolt_spend_pct_of_own_gmv"] = (
        provider_data["avg_daily_bolt_eur"] / provider_data["provider_daily_gmv"]
    )

    provider_data["gmv_quartile"] = pd.qcut(
        provider_data["provider_monthly_gmv"],
        q=4,
        labels=["Q1_small", "Q2_medium", "Q3_large", "Q4_top"],
        duplicates="drop",
    )

    fallback = (
        provider_data.groupby(["gmv_quartile", "Bonus Type"], observed=True)
        .agg(
            avg_bolt_pct_of_own_gmv=("bolt_spend_pct_of_own_gmv", "median"),
            avg_daily_bolt_eur=("avg_daily_bolt_eur", "median"),
            provider_count=("Provider ID", "nunique"),
        )
        .reset_index()
    )

    quartile_bounds = provider_data.groupby("gmv_quartile", observed=True)["provider_monthly_gmv"].agg(
        ["min", "max"]
    )
    print(f"  GMV quartile ranges:")
    for q, row in quartile_bounds.iterrows():
        print(f"    {q}: {row['min']:,.0f} - {row['max']:,.0f} EUR/month")

    return fallback


# ─── 3. Match & Project ─────────────────────────────────────────────────────

def match_and_project(planned_df, benchmarks, fallback, provider_gmv, gmv_df):
    print("\nMatching planned campaigns to historical data...")

    weekly_gmv = {}
    gmv_df_copy = gmv_df.copy()
    gmv_df_copy["date"] = pd.to_datetime(gmv_df_copy["date"])
    for week, (start, end) in WEEK_DATES.items():
        mask = (gmv_df_copy["date"] >= start) & (gmv_df_copy["date"] <= end)
        week_data = gmv_df_copy[mask]
        weekly_gmv[week] = week_data["daily_gmv_eur"].sum()

    results = []
    matched = 0
    fallback_used = 0
    no_match = 0

    if not fallback.empty:
        fallback_quartile_bounds = {}
        for _, row in fallback.iterrows():
            q = row["gmv_quartile"]
            if q not in fallback_quartile_bounds:
                fallback_quartile_bounds[q] = True

    for _, row in planned_df.iterrows():
        provider_id = row["provider_id"]
        bonus_type = row["bonus_type"]
        week = row["week"]

        bm = benchmarks[
            (benchmarks["Provider ID"] == provider_id)
            & (benchmarks["Bonus Type"] == bonus_type)
        ]

        match_method = ""
        avg_daily_eur = 0
        avg_daily_pct = 0

        if len(bm) > 0:
            avg_daily_eur = bm.iloc[0]["avg_daily_bolt_eur"]
            avg_daily_pct = bm.iloc[0]["avg_daily_bolt_pct"]
            match_method = "historical"
            matched += 1
        else:
            bm_any = benchmarks[benchmarks["Provider ID"] == provider_id]
            if len(bm_any) > 0:
                avg_daily_eur = bm_any["avg_daily_bolt_eur"].mean()
                avg_daily_pct = bm_any["avg_daily_bolt_pct"].mean()
                match_method = "historical (diff campaign type)"
                matched += 1
            elif not fallback.empty:
                prov_gmv = provider_gmv.get(provider_id)
                if prov_gmv and prov_gmv > 0:
                    prov_daily_gmv = prov_gmv / 30
                    if prov_gmv <= fallback["avg_daily_bolt_eur"].quantile(0.25):
                        q_label = "Q1_small"
                    elif prov_gmv <= fallback["avg_daily_bolt_eur"].quantile(0.5):
                        q_label = "Q2_medium"
                    elif prov_gmv <= fallback["avg_daily_bolt_eur"].quantile(0.75):
                        q_label = "Q3_large"
                    else:
                        q_label = "Q4_top"

                    fb = fallback[
                        (fallback["Bonus Type"] == bonus_type)
                    ]
                    if len(fb) > 0:
                        avg_bolt_pct = fb["avg_bolt_pct_of_own_gmv"].median()
                        avg_daily_eur = prov_daily_gmv * avg_bolt_pct
                        wk_gmv = weekly_gmv.get(week, 0)
                        avg_daily_pct = avg_daily_eur / (wk_gmv / 7) if wk_gmv > 0 else 0
                        match_method = f"fallback (GMV-based, {prov_gmv:,.0f} EUR/mo)"
                        fallback_used += 1
                    else:
                        fb_all = fallback
                        if len(fb_all) > 0:
                            avg_daily_eur = fb_all["avg_daily_bolt_eur"].median()
                            wk_gmv = weekly_gmv.get(week, 0)
                            avg_daily_pct = avg_daily_eur / (wk_gmv / 7) if wk_gmv > 0 else 0
                            match_method = "fallback (overall median)"
                            fallback_used += 1
                        else:
                            match_method = "NO MATCH"
                            no_match += 1
                else:
                    fb = fallback[fallback["Bonus Type"] == bonus_type]
                    if len(fb) > 0:
                        avg_daily_eur = fb["avg_daily_bolt_eur"].median()
                        wk_gmv = weekly_gmv.get(week, 0)
                        avg_daily_pct = avg_daily_eur / (wk_gmv / 7) if wk_gmv > 0 else 0
                        match_method = "fallback (no provider GMV)"
                        fallback_used += 1
                    else:
                        match_method = "NO MATCH"
                        no_match += 1
            else:
                match_method = "NO MATCH"
                no_match += 1

        weekly_bolt_eur = avg_daily_eur * 7
        weekly_bolt_pct = avg_daily_pct * 7

        results.append({
            **row.to_dict(),
            "match_method": match_method,
            "avg_daily_bolt_eur": avg_daily_eur,
            "avg_daily_bolt_pct_gmv": avg_daily_pct,
            "weekly_projected_bolt_eur": weekly_bolt_eur,
            "weekly_projected_bolt_pct_gmv": weekly_bolt_pct,
        })

    result_df = pd.DataFrame(results)
    print(f"  Historical matches: {matched}")
    print(f"  Fallback matches: {fallback_used}")
    print(f"  No match: {no_match}")
    print(f"  Total projected weekly Bolt spend: {result_df['weekly_projected_bolt_eur'].sum():,.0f} EUR")

    return result_df


# ─── 4. Build Summaries ─────────────────────────────────────────────────────

def build_summary_by_reason(projected_df, gmv_df):
    weekly_gmv = {}
    gmv_df_copy = gmv_df.copy()
    gmv_df_copy["date"] = pd.to_datetime(gmv_df_copy["date"])
    for week, (start, end) in WEEK_DATES.items():
        mask = (gmv_df_copy["date"] >= start) & (gmv_df_copy["date"] <= end)
        weekly_gmv[week] = gmv_df_copy[mask]["daily_gmv_eur"].sum()

    pivot_eur = projected_df.pivot_table(
        index="reason_display",
        columns="week",
        values="weekly_projected_bolt_eur",
        aggfunc="sum",
        fill_value=0,
    )

    for w in TARGET_WEEKS:
        if w not in pivot_eur.columns:
            pivot_eur[w] = 0
    pivot_eur = pivot_eur[TARGET_WEEKS]

    pivot_eur["Total Spend EUR"] = pivot_eur.sum(axis=1)
    pivot_eur[f"Total + {int(ERROR_MARGIN*100)}% margin EUR"] = (
        pivot_eur["Total Spend EUR"] * (1 + ERROR_MARGIN)
    )

    pivot_pct = pivot_eur.copy()
    for w in TARGET_WEEKS:
        wgmv = weekly_gmv.get(w, 1)
        pivot_pct[w] = pivot_eur[w] / wgmv if wgmv > 0 else 0

    total_gmv = sum(weekly_gmv.get(w, 0) for w in TARGET_WEEKS)
    pivot_pct["Total Spend % GMV"] = pivot_eur["Total Spend EUR"] / total_gmv if total_gmv > 0 else 0
    pivot_pct[f"Total + {int(ERROR_MARGIN*100)}% margin % GMV"] = (
        pivot_pct["Total Spend % GMV"] * (1 + ERROR_MARGIN)
    )

    return pivot_eur, pivot_pct, weekly_gmv


def build_summary_by_am(projected_df, gmv_df):
    weekly_gmv = {}
    gmv_df_copy = gmv_df.copy()
    gmv_df_copy["date"] = pd.to_datetime(gmv_df_copy["date"])
    for week, (start, end) in WEEK_DATES.items():
        mask = (gmv_df_copy["date"] >= start) & (gmv_df_copy["date"] <= end)
        weekly_gmv[week] = gmv_df_copy[mask]["daily_gmv_eur"].sum()

    pivot_eur = projected_df.pivot_table(
        index="am_name",
        columns="week",
        values="weekly_projected_bolt_eur",
        aggfunc="sum",
        fill_value=0,
    )
    for w in TARGET_WEEKS:
        if w not in pivot_eur.columns:
            pivot_eur[w] = 0
    pivot_eur = pivot_eur[TARGET_WEEKS]
    pivot_eur["Total Spend EUR"] = pivot_eur.sum(axis=1)
    pivot_eur[f"Total + {int(ERROR_MARGIN*100)}% margin EUR"] = (
        pivot_eur["Total Spend EUR"] * (1 + ERROR_MARGIN)
    )

    return pivot_eur, weekly_gmv


def build_am_reason_breakdown(projected_df):
    breakdown = projected_df.pivot_table(
        index=["am_name", "reason_display"],
        columns="week",
        values="weekly_projected_bolt_eur",
        aggfunc="sum",
        fill_value=0,
    )
    for w in TARGET_WEEKS:
        if w not in breakdown.columns:
            breakdown[w] = 0
    breakdown = breakdown[TARGET_WEEKS]
    breakdown["Total Spend EUR"] = breakdown.sum(axis=1)
    return breakdown


# ─── 5. Write Output Excel ──────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
MARGIN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
MARGIN_FONT = Font(name="Calibri", bold=True, color="BF8F00", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col, eur_cols=None, pct_cols=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if eur_cols and c in eur_cols:
                cell.number_format = '#,##0'
            elif pct_cols and c in pct_cols:
                cell.number_format = '0.000%'


def write_summary_sheet(wb, sheet_name, pivot_eur, pivot_pct, weekly_gmv):
    ws = wb.create_sheet(sheet_name)

    # Title
    ws.cell(row=1, column=1, value="CZ - Spend Projection - March 2026")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    # EUR section
    ws.cell(row=3, column=1, value="Projected Bolt Spend (EUR)")
    ws.cell(row=3, column=1).font = Font(name="Calibri", bold=True, size=12, color="2F5496")

    headers = ["Spend Objective"] + [f"W{w}" for w in TARGET_WEEKS] + ["Total Spend", f"Total + {int(ERROR_MARGIN*100)}% margin"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header(ws, 4, len(headers))

    row_idx = 5
    for reason in sorted(pivot_eur.index):
        ws.cell(row=row_idx, column=1, value=reason)
        for ci, w in enumerate(TARGET_WEEKS, 2):
            ws.cell(row=row_idx, column=ci, value=pivot_eur.loc[reason, w])
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2, value=pivot_eur.loc[reason, "Total Spend EUR"])
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 3, value=pivot_eur.loc[reason, f"Total + {int(ERROR_MARGIN*100)}% margin EUR"])
        row_idx += 1

    # Total row
    total_row = row_idx
    ws.cell(row=total_row, column=1, value="Total")
    for ci, w in enumerate(TARGET_WEEKS, 2):
        ws.cell(row=total_row, column=ci, value=pivot_eur[w].sum())
    ws.cell(row=total_row, column=len(TARGET_WEEKS) + 2, value=pivot_eur["Total Spend EUR"].sum())
    ws.cell(row=total_row, column=len(TARGET_WEEKS) + 3, value=pivot_eur[f"Total + {int(ERROR_MARGIN*100)}% margin EUR"].sum())
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_row, column=c).font = TOTAL_FONT
        ws.cell(row=total_row, column=c).border = THIN_BORDER
    row_idx += 1

    # Margin row
    margin_row = row_idx
    ws.cell(row=margin_row, column=1, value=f"Total + {int(ERROR_MARGIN*100)}% error margin")
    for ci, w in enumerate(TARGET_WEEKS, 2):
        ws.cell(row=margin_row, column=ci, value=pivot_eur[w].sum() * (1 + ERROR_MARGIN))
    for c in range(1, len(headers) + 1):
        ws.cell(row=margin_row, column=c).fill = MARGIN_FILL
        ws.cell(row=margin_row, column=c).font = MARGIN_FONT
        ws.cell(row=margin_row, column=c).border = THIN_BORDER
    row_idx += 1

    eur_cols = set(range(2, len(headers) + 1))
    style_body(ws, 5, total_row - 1, len(headers), eur_cols=eur_cols)

    # % of GMV section
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="Projected Bolt Spend (% of GMV)")
    ws.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    row_idx += 1

    headers_pct = ["Spend Objective"] + [f"W{w}" for w in TARGET_WEEKS] + ["Total % of GMV", f"Total + {int(ERROR_MARGIN*100)}% margin"]
    for ci, h in enumerate(headers_pct, 1):
        ws.cell(row=row_idx, column=ci, value=h)
    style_header(ws, row_idx, len(headers_pct))
    row_idx += 1

    pct_start = row_idx
    total_gmv = sum(weekly_gmv.get(w, 0) for w in TARGET_WEEKS)
    for reason in sorted(pivot_eur.index):
        ws.cell(row=row_idx, column=1, value=reason)
        for ci, w in enumerate(TARGET_WEEKS, 2):
            wgmv = weekly_gmv.get(w, 1)
            ws.cell(row=row_idx, column=ci, value=pivot_eur.loc[reason, w] / wgmv if wgmv > 0 else 0)
        total_pct = pivot_eur.loc[reason, "Total Spend EUR"] / total_gmv if total_gmv > 0 else 0
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2, value=total_pct)
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 3, value=total_pct * (1 + ERROR_MARGIN))
        row_idx += 1

    total_pct_row = row_idx
    ws.cell(row=total_pct_row, column=1, value="Total")
    for ci, w in enumerate(TARGET_WEEKS, 2):
        wgmv = weekly_gmv.get(w, 1)
        ws.cell(row=total_pct_row, column=ci, value=pivot_eur[w].sum() / wgmv if wgmv > 0 else 0)
    ws.cell(row=total_pct_row, column=len(TARGET_WEEKS) + 2,
            value=pivot_eur["Total Spend EUR"].sum() / total_gmv if total_gmv > 0 else 0)
    for c in range(1, len(headers_pct) + 1):
        ws.cell(row=total_pct_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_pct_row, column=c).font = TOTAL_FONT
        ws.cell(row=total_pct_row, column=c).border = THIN_BORDER

    pct_cols = set(range(2, len(headers_pct) + 1))
    style_body(ws, pct_start, total_pct_row - 1, len(headers_pct), pct_cols=pct_cols)

    # Weekly GMV reference
    row_idx = total_pct_row + 3
    ws.cell(row=row_idx, column=1, value="Reference: Weekly GMV (EUR)")
    ws.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, italic=True, size=10, color="808080")
    row_idx += 1
    for ci, w in enumerate(TARGET_WEEKS, 2):
        ws.cell(row=row_idx, column=ci - 1, value=f"W{w}")
        ws.cell(row=row_idx, column=ci, value=weekly_gmv.get(w, 0))
        ws.cell(row=row_idx, column=ci).number_format = '#,##0'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 28


def write_am_summary_sheet(wb, projected_df, weekly_gmv):
    ws = wb.create_sheet("Summary per AM")

    ws.cell(row=1, column=1, value="CZ - Spend Projection by AM - March 2026")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    headers = ["Account Manager"] + [f"W{w}" for w in TARGET_WEEKS] + ["Total Spend EUR", f"Total + {int(ERROR_MARGIN*100)}% margin EUR"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, len(headers))

    row_idx = 4
    am_totals = projected_df.groupby("am_name").agg(
        total=("weekly_projected_bolt_eur", "sum")
    ).sort_values("total", ascending=False)

    total_gmv = sum(weekly_gmv.get(w, 0) for w in TARGET_WEEKS)

    for am_name in am_totals.index:
        am_data = projected_df[projected_df["am_name"] == am_name]

        # AM header row
        ws.cell(row=row_idx, column=1, value=am_name)
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, size=11)
        am_week_totals = am_data.groupby("week")["weekly_projected_bolt_eur"].sum()
        total_am = 0
        for ci, w in enumerate(TARGET_WEEKS, 2):
            val = am_week_totals.get(w, 0)
            ws.cell(row=row_idx, column=ci, value=val)
            ws.cell(row=row_idx, column=ci).number_format = '#,##0'
            ws.cell(row=row_idx, column=ci).font = Font(name="Calibri", bold=True, size=11)
            total_am += val
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2, value=total_am)
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2).number_format = '#,##0'
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2).font = TOTAL_FONT
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 3, value=total_am * (1 + ERROR_MARGIN))
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 3).number_format = '#,##0'
        ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 3).font = MARGIN_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).fill = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            )
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
        row_idx += 1

        # Breakdown by reason
        reason_breakdown = am_data.groupby(["reason_display", "week"])["weekly_projected_bolt_eur"].sum().unstack(fill_value=0)
        for reason in sorted(reason_breakdown.index):
            ws.cell(row=row_idx, column=1, value=f"  {reason}")
            ws.cell(row=row_idx, column=1).font = Font(name="Calibri", size=10, color="595959")
            reason_total = 0
            for ci, w in enumerate(TARGET_WEEKS, 2):
                val = reason_breakdown.loc[reason].get(w, 0)
                ws.cell(row=row_idx, column=ci, value=val)
                ws.cell(row=row_idx, column=ci).number_format = '#,##0'
                ws.cell(row=row_idx, column=ci).font = Font(name="Calibri", size=10, color="595959")
                ws.cell(row=row_idx, column=ci).border = THIN_BORDER
                reason_total += val
            ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2, value=reason_total)
            ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2).number_format = '#,##0'
            ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2).font = Font(name="Calibri", size=10, color="595959")
            ws.cell(row=row_idx, column=len(TARGET_WEEKS) + 2).border = THIN_BORDER
            row_idx += 1

        row_idx += 1

    # Grand Total
    grand_total_row = row_idx
    ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL")
    grand_week_totals = projected_df.groupby("week")["weekly_projected_bolt_eur"].sum()
    grand_total = 0
    for ci, w in enumerate(TARGET_WEEKS, 2):
        val = grand_week_totals.get(w, 0)
        ws.cell(row=grand_total_row, column=ci, value=val)
        grand_total += val
    ws.cell(row=grand_total_row, column=len(TARGET_WEEKS) + 2, value=grand_total)
    ws.cell(row=grand_total_row, column=len(TARGET_WEEKS) + 3, value=grand_total * (1 + ERROR_MARGIN))

    for c in range(1, len(headers) + 1):
        ws.cell(row=grand_total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=grand_total_row, column=c).font = TOTAL_FONT
        ws.cell(row=grand_total_row, column=c).border = THIN_BORDER
        ws.cell(row=grand_total_row, column=c).number_format = '#,##0'

    # % of GMV row
    pct_row = grand_total_row + 1
    ws.cell(row=pct_row, column=1, value="% of GMV")
    for ci, w in enumerate(TARGET_WEEKS, 2):
        wgmv = weekly_gmv.get(w, 1)
        ws.cell(row=pct_row, column=ci, value=grand_week_totals.get(w, 0) / wgmv if wgmv > 0 else 0)
        ws.cell(row=pct_row, column=ci).number_format = '0.000%'
    ws.cell(row=pct_row, column=len(TARGET_WEEKS) + 2,
            value=grand_total / total_gmv if total_gmv > 0 else 0)
    ws.cell(row=pct_row, column=len(TARGET_WEEKS) + 2).number_format = '0.000%'
    for c in range(1, len(headers) + 1):
        ws.cell(row=pct_row, column=c).font = Font(name="Calibri", italic=True, size=10, color="808080")
        ws.cell(row=pct_row, column=c).border = THIN_BORDER

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 30


def write_detail_sheet(wb, projected_df):
    ws = wb.create_sheet("Campaign Detail")

    ws.cell(row=1, column=1, value="Campaign-Level Projection Detail")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    detail_cols = [
        "am_name", "week", "provider_id", "provider_name", "city",
        "campaign_type", "discount_pct", "cost_share_pct",
        "reason_display", "commitment", "match_method",
        "avg_daily_bolt_eur", "weekly_projected_bolt_eur",
    ]
    display_names = [
        "Account Manager", "Week", "Provider ID", "Provider Name", "City",
        "Campaign Type", "Discount %", "Cost Share % (Provider)", 
        "Spend Objective", "Commitment", "Match Method",
        "Avg Daily Bolt Spend (EUR)", "Weekly Projected Bolt Spend (EUR)",
    ]

    for ci, name in enumerate(display_names, 1):
        ws.cell(row=3, column=ci, value=name)
    style_header(ws, 3, len(display_names))

    sorted_df = projected_df.sort_values(
        ["am_name", "week", "weekly_projected_bolt_eur"],
        ascending=[True, True, False],
    )

    for ri, (_, row) in enumerate(sorted_df.iterrows(), 4):
        for ci, col in enumerate(detail_cols, 1):
            val = row[col]
            ws.cell(row=ri, column=ci, value=val)
            ws.cell(row=ri, column=ci).font = BODY_FONT
            ws.cell(row=ri, column=ci).border = THIN_BORDER

            if col in ("avg_daily_bolt_eur", "weekly_projected_bolt_eur"):
                ws.cell(row=ri, column=ci).number_format = '#,##0.00'
            elif col == "discount_pct":
                ws.cell(row=ri, column=ci).number_format = '0'

        if "fallback" in str(row["match_method"]).lower() or "NO MATCH" in str(row["match_method"]):
            for ci in range(1, len(detail_cols) + 1):
                ws.cell(row=ri, column=ci).fill = PatternFill(
                    start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"
                )

    col_widths = [22, 7, 12, 30, 20, 16, 10, 18, 20, 12, 35, 22, 28]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_output(projected_df, pivot_eur_reason, pivot_pct_reason, weekly_gmv):
    print(f"\nWriting output to {OUTPUT_FILE}...")
    wb = openpyxl.Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    write_summary_sheet(wb, "Summary", pivot_eur_reason, pivot_pct_reason, weekly_gmv)
    write_am_summary_sheet(wb, projected_df, weekly_gmv)
    write_detail_sheet(wb, projected_df)

    wb.save(OUTPUT_FILE)
    print(f"  Done! Saved to {OUTPUT_FILE}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("CZ Spend Projection Tool - March 2026")
    print("=" * 60)

    hist_df = read_historical_csv()
    gmv_df = read_daily_gmv()
    planned_df = read_am_files()
    provider_gmv = read_provider_gmv_from_data_drop()

    benchmarks = build_historical_benchmarks(hist_df, gmv_df)
    fallback = build_fallback_benchmarks(hist_df, provider_gmv, gmv_df)

    projected_df = match_and_project(planned_df, benchmarks, fallback, provider_gmv, gmv_df)

    pivot_eur_reason, pivot_pct_reason, weekly_gmv = build_summary_by_reason(projected_df, gmv_df)

    print("\n── Summary by Spend Objective (EUR) ──")
    print(pivot_eur_reason.to_string(float_format=lambda x: f"{x:,.0f}"))

    print("\n── Summary by AM (EUR) ──")
    am_summary = projected_df.groupby("am_name")["weekly_projected_bolt_eur"].sum().sort_values(ascending=False)
    print(am_summary.to_string(float_format=lambda x: f"{x:,.0f}"))

    write_output(projected_df, pivot_eur_reason, pivot_pct_reason, weekly_gmv)

    total_spend = projected_df["weekly_projected_bolt_eur"].sum()
    total_gmv = sum(weekly_gmv.get(w, 0) for w in TARGET_WEEKS)
    print(f"\n{'=' * 60}")
    print(f"TOTAL PROJECTED BOLT SPEND (March): {total_spend:,.0f} EUR")
    print(f"TOTAL PROJECTED GMV (March): {total_gmv:,.0f} EUR")
    print(f"SPEND AS % OF GMV: {total_spend / total_gmv * 100:.3f}%")
    print(f"SPEND + {int(ERROR_MARGIN*100)}% MARGIN: {total_spend * (1 + ERROR_MARGIN):,.0f} EUR")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
